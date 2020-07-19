import os
print("working directory:", os.getcwd())
from Bio import Entrez, Medline
import requests
import openpyxl
import time
from tqdm import trange
from bs4 import BeautifulSoup

class pubmed_utils():
    def __init__(self):
        pass
        
        
    def get_main_info_into_excel(self, email, search_key_words, release_date_cutoff, paper_type, grab_total, save_path):
        '''
        grab info from pubmed, save it into a excel
        '''
        
        Entrez.email = email
        Entrez.tool = "MyLocalScript"
        grab_step = 10
        if grab_total==None:
            grab_total = total
        
        esearch_require = Entrez.esearch(db="pubmed", term=search_key_words, reldate=release_date_cutoff, ptyp=paper_type, usehistory="y")
        read_esearch = Entrez.read(esearch_require)
        total = int(read_esearch["Count"])
        print("Find total:", total)
        webenv = read_esearch["WebEnv"]
        query_key = read_esearch["QueryKey"]

        self.excel_property_dic = {token:index for index, token in enumerate(["PMID", "TI", "TA", "IF", "LR", "AB", "LID"], start=1)}
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=self.excel_property_dic["PMID"]).value = "PMID"
        ws.cell(row=1, column=self.excel_property_dic["TI"]).value = "Title"
        ws.cell(row=1, column=self.excel_property_dic["TA"]).value = "Journal"
        ws.cell(row=1, column=self.excel_property_dic["IF"]).value = "IF"
        ws.cell(row=1, column=self.excel_property_dic["LR"]).value = "publish_date"
        ws.cell(row=1, column=self.excel_property_dic["AB"]).value = "Abstract"
        ws.cell(row=1, column=self.excel_property_dic["LID"]).value = "DOI"

        cur_row = 2
        for step in trange(0, grab_total//grab_step, desc="getting pubmed info"):
            efetch_require = Entrez.efetch(db="pubmed", retstart=step*grab_step, retmax=grab_step, webenv=webenv, query_key=query_key, rettype="medline", retmode="text")
            records = Medline.parse(efetch_require)
            records = list(records)
            
            for record in records:
                for key in self.excel_property_dic.keys():
                    if key not in record.keys():
                        continue
                    key_info = record[key]
                    ws.cell(row=cur_row, column=self.excel_property_dic[key]).value = key_info
                cur_row += 1

        wb.save(save_path)
        
        
    def embed_IF_into_excel(self, excel_path):
        '''
        grab IF from scholarscope and save it into excel
        '''
        
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Sheet"]
        j_list = []
        for cur_row in range(2, ws.max_row+1):
            j_name = ws.cell(row=cur_row, column=self.excel_property_dic["TA"]).value
            if j_name not in j_list:
                j_list.append(j_name)
        
        IF_url = "https://api.scholarscope.cn/getsinglesearch.php"
        session_requests = requests.session()
        j_dic = {}
        fail_list = []
        for i in trange(len(j_list), desc="getting IF info"):
            headers = {'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4128.3 Safari/537.36'}
            data = {'jrnl': j_list[i]}
            result = session_requests.post(IF_url, data=data, headers=headers)
            result = str(result.content, "utf-8")
            result_soup = BeautifulSoup(result, "html.parser")
            try:
                IF = float(result_soup.table.findAll("tr")[1].findAll("td")[2].string)
                j_dic[j_list[i]]=IF
            except:
                j_dic[j_list[i]]="Unknow"
                fail_list.append(j_list[i])
            time.sleep(1)                 
        print("failure total:", len(fail_list))
        print("failure list:", fail_list)
        
        for cur_row in range(2, ws.max_row+1):
            j_name = ws.cell(row=cur_row, column=self.excel_property_dic["TA"]).value
            ws.cell(row=cur_row, column=self.excel_property_dic["IF"]).value = j_dic[j_name]
        wb.save(excel_path)
    
    
    def download_pdf(self, excel_path, pdf_savepath, IF_cutoff):
        '''
        try to download paper which IF higher than cutoff
        warning: very low successful rate
        '''
        
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Sheet"]
        base_url = "https://sci-hub.tw/"
        success_count = 0
        for cur_row in trange(2, ws.max_row+1, desc="downloading pdf"):
            IF = ws.cell(row=cur_row, column=self.excel_property_dic["IF"]).value
            pmid = ws.cell(row=cur_row, column=self.excel_property_dic["PMID"]).value
            title = ws.cell(row=cur_row, column=self.excel_property_dic["TI"]).value
            if IF=="Unknow" or float(IF)<IF_cutoff:
                continue

            file_name = pdf_savepath+pmid+"_"+title+".pdf"
            try:
                doi = ws.cell(row=cur_row, column=self.excel_property_dic["LID"]).value.split(" ")[0]
                url = base_url + doi
                getpage = requests.get(url, verify=True)
                getpage_soup = BeautifulSoup(getpage.text, "html.parser")
                src = getpage_soup.find("iframe", src=True).get_attribute_list("src")[0]
                response = requests.get("https:"+src, verify=True)
                f = open(file_name, "wb+")
                f.write(response.content)
                f.close()
                success_count += 1
            except:
                pass
            time.sleep(1)
        print("successful download: {}".format(success_count))